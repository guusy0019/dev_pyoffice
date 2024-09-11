import unittest
from openpyxl import Workbook

from app.module.utility.xlsx_utility import XlsxUtility


class TestXlsxUtility(unittest.TestCase):

    def setUp(self):
        self.xlsx_utility = XlsxUtility()
        self.xlsx_path = "C:\\Users\\matty\\doc\\home_doc\\2023（sample）出勤簿.xlsx"
        self.xlsx_path_key = "2023（sample）出勤簿.xlsx"

    def test_load_workbook(self):
        workbook: Workbook = self.xlsx_utility.load_workbook(self.xlsx_path)
        self.assertIsNotNone(workbook)

    def test_read_sheet(self):
        sheet_name = "sample_sheet"
        sheet = self.xlsx_utility.read_sheet(self.xlsx_path, sheet_name)
        self.assertIsNotNone(sheet)

    def test_save_workbook(self):
        workbook: Workbook = self.xlsx_utility.save_workbook(self.xlsx_path)
        self.assertIsNotNone(workbook)
        # ファイルを閉じないと、他のプロセスでこのファイルに書き込む際にエラーが発生する「permission denied」
        workbook.close()
        return None

    def test_save_worksheet(self):
        sheet_name = "sample_sheet_by_test"
        workbook: Workbook = self.xlsx_utility.save_workbook(self.xlsx_path)
        sheet = workbook.create_sheet(sheet_name)
        # saveしないとワークシートを保存しないっぽい
        workbook.save(self.xlsx_path)
        workbook.close()
        return None
