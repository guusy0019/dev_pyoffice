import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class XlsxUtility:

    @staticmethod
    def load_workbook(xlsx_path: str) -> Workbook:
        return openpyxl.load_workbook(xlsx_path)

    @staticmethod
    def read_sheet(xlsx_path: str, sheet_name: str) -> Worksheet:
        workbook = openpyxl.load_workbook(xlsx_path)
        return workbook[sheet_name]

    @staticmethod
    def save_workbook(xlsx_path: str) -> None:
        Workbook = openpyxl.Workbook()
        Workbook.save(xlsx_path)
        return Workbook

    @staticmethod
    def save_worksheet(xlsx_path: str, sheet_name: str) -> None:
        workbook = openpyxl.load_workbook(xlsx_path)
        workbook.create_sheet(sheet_name)
        workbook.save(xlsx_path)
        return None
